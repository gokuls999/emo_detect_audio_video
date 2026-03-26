"""SQLite persistence for MSTDN-A Dashboard."""
from __future__ import annotations
import json, sqlite3
from datetime import datetime
from pathlib import Path

DB_PATH = Path(__file__).parent / "dashboard.db"

# ── connection ──────────────────────────────────────────────────────────────
def _conn() -> sqlite3.Connection:
    c = sqlite3.connect(DB_PATH, check_same_thread=False)
    c.row_factory = sqlite3.Row
    return c

def init_db() -> None:
    with _conn() as c:
        c.executescript("""
        CREATE TABLE IF NOT EXISTS sessions (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            name        TEXT    NOT NULL,
            location    TEXT    DEFAULT '',
            started_at  TEXT,
            ended_at    TEXT
        );
        CREATE TABLE IF NOT EXISTS participants (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            session_id  INTEGER NOT NULL,
            name        TEXT    NOT NULL,
            role        TEXT    DEFAULT '',
            department  TEXT    DEFAULT '',
            FOREIGN KEY(session_id) REFERENCES sessions(id)
        );
        CREATE TABLE IF NOT EXISTS readings (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            session_id      INTEGER NOT NULL,
            participant_id  INTEGER,
            timestamp       TEXT,
            emotion         TEXT,
            confidence      REAL,
            stress          REAL,
            valence         REAL,
            arousal         REAL,
            probs           TEXT,
            FOREIGN KEY(session_id)     REFERENCES sessions(id),
            FOREIGN KEY(participant_id) REFERENCES participants(id)
        );
        CREATE TABLE IF NOT EXISTS alerts (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            session_id      INTEGER NOT NULL,
            participant_id  INTEGER,
            timestamp       TEXT,
            alert_type      TEXT,
            message         TEXT,
            FOREIGN KEY(session_id) REFERENCES sessions(id)
        );
        CREATE TABLE IF NOT EXISTS face_embeddings (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            participant_id  INTEGER NOT NULL UNIQUE,
            embedding       BLOB    NOT NULL,
            registered_at   TEXT,
            FOREIGN KEY(participant_id) REFERENCES participants(id)
        );
        """)

# ── sessions ────────────────────────────────────────────────────────────────
def create_session(name: str, location: str = "") -> int:
    with _conn() as c:
        cur = c.execute(
            "INSERT INTO sessions (name, location, started_at) VALUES (?,?,?)",
            (name, location, datetime.now().isoformat())
        )
        return cur.lastrowid

def end_session(session_id: int) -> None:
    with _conn() as c:
        c.execute("UPDATE sessions SET ended_at=? WHERE id=?",
                  (datetime.now().isoformat(), session_id))

def list_sessions() -> list[dict]:
    with _conn() as c:
        rows = c.execute("SELECT * FROM sessions ORDER BY id DESC").fetchall()
        return [dict(r) for r in rows]

def get_session(session_id: int) -> dict | None:
    with _conn() as c:
        row = c.execute("SELECT * FROM sessions WHERE id=?", (session_id,)).fetchone()
        return dict(row) if row else None

# ── participants ─────────────────────────────────────────────────────────────
def add_participant(session_id: int, name: str, role: str = "", department: str = "") -> int:
    with _conn() as c:
        cur = c.execute(
            "INSERT INTO participants (session_id, name, role, department) VALUES (?,?,?,?)",
            (session_id, name, role, department)
        )
        return cur.lastrowid

def remove_participant(participant_id: int) -> None:
    with _conn() as c:
        c.execute("DELETE FROM participants WHERE id=?", (participant_id,))

def get_participants(session_id: int) -> list[dict]:
    with _conn() as c:
        rows = c.execute(
            "SELECT * FROM participants WHERE session_id=?", (session_id,)
        ).fetchall()
        return [dict(r) for r in rows]

# ── readings ─────────────────────────────────────────────────────────────────
def log_reading(session_id: int, participant_id: int | None,
                emotion: str, confidence: float, stress: float,
                valence: float, arousal: float, probs: list) -> int:
    with _conn() as c:
        cur = c.execute(
            """INSERT INTO readings
               (session_id, participant_id, timestamp, emotion,
                confidence, stress, valence, arousal, probs)
               VALUES (?,?,?,?,?,?,?,?,?)""",
            (session_id, participant_id, datetime.now().isoformat(),
             emotion, confidence, stress, valence, arousal, json.dumps(probs))
        )
        return cur.lastrowid

def get_readings(session_id: int, participant_id: int | None = None,
                 limit: int = 200) -> list[dict]:
    with _conn() as c:
        if participant_id is not None:
            rows = c.execute(
                "SELECT * FROM readings WHERE session_id=? AND participant_id=? "
                "ORDER BY id DESC LIMIT ?", (session_id, participant_id, limit)
            ).fetchall()
        else:
            rows = c.execute(
                "SELECT * FROM readings WHERE session_id=? ORDER BY id DESC LIMIT ?",
                (session_id, limit)
            ).fetchall()
        return [dict(r) for r in rows]

def session_summary(session_id: int) -> dict:
    with _conn() as c:
        rows = c.execute(
            "SELECT emotion, stress, valence, arousal FROM readings WHERE session_id=?",
            (session_id,)
        ).fetchall()
        if not rows:
            return {}
        import statistics
        stresses = [r["stress"] for r in rows]
        valences = [r["valence"] for r in rows]
        arousals = [r["arousal"] for r in rows]
        from collections import Counter
        emotions = Counter(r["emotion"] for r in rows)
        return {
            "total_readings":      len(rows),
            "dominant_emotion":    emotions.most_common(1)[0][0],
            "emotion_counts":      dict(emotions),
            "mean_stress":         round(statistics.mean(stresses), 3),
            "max_stress":          round(max(stresses), 3),
            "mean_valence":        round(statistics.mean(valences), 3),
            "mean_arousal":        round(statistics.mean(arousals), 3),
            "high_stress_pct":     round(sum(s > 0.7 for s in stresses) / len(stresses) * 100, 1),
        }

# ── alerts ───────────────────────────────────────────────────────────────────
def log_alert(session_id: int, participant_id: int | None,
              alert_type: str, message: str) -> None:
    with _conn() as c:
        c.execute(
            "INSERT INTO alerts (session_id, participant_id, timestamp, alert_type, message) "
            "VALUES (?,?,?,?,?)",
            (session_id, participant_id, datetime.now().isoformat(), alert_type, message)
        )

def get_alerts(session_id: int) -> list[dict]:
    with _conn() as c:
        rows = c.execute(
            "SELECT * FROM alerts WHERE session_id=? ORDER BY id DESC",
            (session_id,)
        ).fetchall()
        return [dict(r) for r in rows]

# ── export ───────────────────────────────────────────────────────────────────
def export_session_excel(session_id: int, out_path: str) -> None:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = openpyxl.Workbook()

    # Summary sheet
    ws = wb.active
    ws.title = "Summary"
    summary = session_summary(session_id)
    session = get_session(session_id)
    ws.append(["MSTDN-A Session Report"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append(["Session", session["name"]])
    ws.append(["Location", session.get("location", "")])
    ws.append(["Started", session["started_at"]])
    ws.append(["Ended", session.get("ended_at", "ongoing")])
    ws.append([])
    for k, v in summary.items():
        ws.append([k.replace("_", " ").title(), str(v)])

    # Readings sheet
    ws2 = wb.create_sheet("Readings")
    headers = ["Timestamp", "Participant", "Emotion", "Confidence",
               "Stress", "Valence", "Arousal"]
    ws2.append(headers)
    for cell in ws2[1]:
        cell.font = Font(bold=True)

    participants = {p["id"]: p["name"] for p in get_participants(session_id)}
    readings = get_readings(session_id, limit=10000)
    for r in reversed(readings):
        ws2.append([
            r["timestamp"],
            participants.get(r["participant_id"], "Group"),
            r["emotion"],
            round(r["confidence"], 3),
            round(r["stress"], 3),
            round(r["valence"], 3),
            round(r["arousal"], 3),
        ])

    # Alerts sheet
    ws3 = wb.create_sheet("Alerts")
    ws3.append(["Timestamp", "Type", "Message"])
    for a in get_alerts(session_id):
        ws3.append([a["timestamp"], a["alert_type"], a["message"]])

    # Per-participant sheet
    for p in get_participants(session_id):
        ws_p = wb.create_sheet(p["name"][:25])
        ws_p.append(["Timestamp","Emotion","Confidence","Stress","Valence","Arousal"])
        for cell in ws_p[1]: cell.font = Font(bold=True)
        for r in reversed(get_readings(session_id, p["id"], limit=10000)):
            ws_p.append([r["timestamp"], r["emotion"],
                         round(r["confidence"],3), round(r["stress"],3),
                         round(r["valence"],3),   round(r["arousal"],3)])

    wb.save(out_path)

# ── face embeddings ────────────────────────────────────────────────────────────
def store_face_embedding(participant_id: int, embedding: "np.ndarray") -> None:
    import numpy as np
    blob = embedding.astype(np.float32).tobytes()
    with _conn() as c:
        c.execute("""INSERT INTO face_embeddings (participant_id, embedding, registered_at)
                     VALUES (?,?,?)
                     ON CONFLICT(participant_id) DO UPDATE SET
                       embedding=excluded.embedding,
                       registered_at=excluded.registered_at""",
                  (participant_id, blob, datetime.now().isoformat()))

def get_face_embeddings() -> list[dict]:
    """Return all registered face embeddings as list of {participant_id, name, embedding}."""
    import numpy as np
    with _conn() as c:
        rows = c.execute("""
            SELECT fe.participant_id, p.name, fe.embedding
            FROM face_embeddings fe
            JOIN participants p ON p.id = fe.participant_id
        """).fetchall()
    result = []
    for r in rows:
        emb = np.frombuffer(r["embedding"], dtype=np.float32).copy()
        result.append({"participant_id": r["participant_id"],
                       "name": r["name"], "embedding": emb})
    return result

def has_face(participant_id: int) -> bool:
    with _conn() as c:
        row = c.execute("SELECT id FROM face_embeddings WHERE participant_id=?",
                        (participant_id,)).fetchone()
        return row is not None

init_db()
